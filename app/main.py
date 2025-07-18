from fastapi import FastAPI, Request, Depends, HTTPException, UploadFile, File
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse, Response, FileResponse
from typing import List, Optional
from sqlalchemy.orm import Session
from pydantic import BaseModel
from datetime import datetime

# Importações do projeto
from app.models import IndividuoCreate, ResultadoProcessamentoIndividual, ErroLinha
from app.services.anthropometry_service import AnthropometryService
from app.db.session import get_db
from app.core.config import settings

from fastapi.staticfiles import StaticFiles
from app.models import ResultadoProcessamentoIndividual, ErroLinha

class BatchProcessingResponse(BaseModel):
    """Resposta padronizada para processamento em lote"""
    success: bool
    summary: dict
    results: List[ResultadoProcessamentoIndividual]
    errors: List[ErroLinha]

# Importação para o novo gerador de PDF (Puro Python)
from fpdf import FPDF
from fpdf.enums import Align

app = FastAPI(title=settings.APP_NAME, debug=settings.DEBUG)

app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")

@app.get("/", response_class=HTMLResponse, include_in_schema=False)
async def read_root(request: Request):
    """Serve a página principal da aplicação."""
    return templates.TemplateResponse(request, "index.html", {"title": "Avaliação Antropométrica"})

@app.get("/favicon.ico", include_in_schema=False)
async def favicon():
    """Serve o favicon da aplicação."""
    return FileResponse("app/static/favicon.svg", media_type="image/svg+xml")

@app.post("/api/processar/individual", response_model=ResultadoProcessamentoIndividual)
async def processar_dados_individuais(
    individuo_data: IndividuoCreate,
    db: Session = Depends(get_db)
):
    """Processa os dados de um único indivíduo e retorna a avaliação nutricional."""
    try:
        service = AnthropometryService(db=db)
        return service.process_individual_data(individuo_data)
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        print(f"Erro inesperado no processamento individual: {e}")
        raise HTTPException(status_code=500, detail=f"Erro interno no servidor: {e}")

@app.post("/api/processar/lote", response_model=BatchProcessingResponse)
async def processar_dados_lote(
    batchFile: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Processa um arquivo em lote (CSV/TSV) e retorna os resultados."""
    if not batchFile.filename:
        raise HTTPException(status_code=400, detail="Nenhum arquivo foi enviado.")
    
    if not (batchFile.filename.endswith(".csv") or batchFile.filename.endswith(".tsv")):
        raise HTTPException(status_code=400, detail="Formato de arquivo inválido. Use CSV ou TSV.")
    
    try:
        contents = await batchFile.read()
        if not contents:
            raise HTTPException(status_code=400, detail="Arquivo está vazio.")
            
        service = AnthropometryService(db=db)
        processed_data = service.process_batch_data(contents, batchFile.filename)
        
        return BatchProcessingResponse(
            success=True,
            summary={
                "total_processed": processed_data["total_rows_attempted"],
                "success_count": len(processed_data["resultados_individuais"]),
                "error_count": len(processed_data["erros_por_linha"])
            },
            results=processed_data["resultados_individuais"],
            errors=processed_data["erros_por_linha"]
        )
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        print(f"Erro inesperado no processamento em lote: {e}")
        raise HTTPException(status_code=500, detail=f"Erro interno no servidor ao processar o lote: {e}")


class ReportData(BaseModel):
    """Modelo de dados para receber as informações para o relatório PDF."""
    identifier: str
    sub_identifier: Optional[str] = None
    batch_results: BatchProcessingResponse

class PDF(FPDF):
    """Classe customizada para criar o PDF com cabeçalho e rodapé."""
    def __init__(self, identifier: str = '', sub_identifier: str = ''):
        super().__init__(orientation='L', unit='mm', format='A4')
        self.identifier = identifier
        self.sub_identifier = sub_identifier
        self.set_auto_page_break(auto=True, margin=15)

    def header(self):
        self.set_font('Helvetica', 'B', 16)
        self.cell(0, 10, self.identifier, 0, 1, 'L')
        if self.sub_identifier:
            self.set_font('Helvetica', '', 12)
            self.cell(0, 8, self.sub_identifier, 0, 1, 'L')
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font('Helvetica', 'I', 8)
        now = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        self.cell(0, 10, f'Página {self.page_no()}/{{nb}}', 0, 0, 'C')
        self.cell(0, 10, f'Gerado em: {now}', 0, 0, 'R')

def get_fpdf_color(classification_text: Optional[str]):
    """Retorna uma tupla de cor RGB para o fpdf2 com base no texto da classificação."""
    if not classification_text:
        return None
    lower_class = classification_text.lower().replace(' ', '_')
    if any(s in lower_class for s in ['magreza_acentuada', 'muito_baixo_peso', 'obesidade_grave', 'erro']): return (254, 226, 226)
    if any(s in lower_class for s in ['magreza', 'baixo_peso', 'baixa_estatura', 'obesidade']): return (254, 242, 242)
    if any(s in lower_class for s in ['eutrofia', 'peso_adequado', 'estatura_adequada']): return (236, 253, 245)
    if any(s in lower_class for s in ['risco_de_sobrepeso', 'sobrepeso']): return (254, 252, 232)
    return None

def encode_for_latin1(text: str) -> str:
    """Codifica o texto para o formato latin-1, substituindo caracteres incompatíveis."""
    return text.encode('latin-1', 'replace').decode('latin-1')

@app.post("/api/export/pdf")
async def export_pdf_fpdf(report_data: ReportData):
    """Gera um relatório PDF usando FPDF2."""
    try:
        pdf = PDF(identifier=report_data.identifier, sub_identifier=report_data.sub_identifier or '')
        pdf.alias_nb_pages()
        pdf.add_page()
        
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_fill_color(22, 78, 99)
        pdf.set_text_color(255, 255, 255)
        col_widths = (25, 35, 25, 20, 20, 17, 35, 17, 38, 10, 30)
        header = ['ID', 'Nome', 'Idade', 'Data Nasc.', 'Data Aval.', 'Peso(kg)', 'P/I', 'Altura(cm)', 'A/I', 'IMC', 'IMC/I']
        for i, header_text in enumerate(header):
            pdf.cell(col_widths[i], 8, header_text, border=1, align='C', fill=True)
        pdf.ln()

        pdf.set_font('Helvetica', '', 7)
        pdf.set_text_color(0, 0, 0)
        
        def get_classification_from_indicators(indicators, prefix):
            return next((i.classificacao for i in indicators if i.tipo.startswith(prefix)), "N/A")

        for res in report_data.batch_results.results or []:
            imc_i_class = get_classification_from_indicators(res.indicadores, "IMC-para-Idade")
            color = get_fpdf_color(imc_i_class)
            fill = bool(color)
            if color:
                pdf.set_fill_color(*color)

            pdf.cell(col_widths[0], 6, encode_for_latin1(res.id_paciente or ''), border=1, align='L', fill=fill)
            pdf.cell(col_widths[1], 6, encode_for_latin1(res.nome or 'N/A'), border=1, align='L', fill=fill)
            pdf.cell(col_widths[2], 6, res.idade if res.idade else "N/A", border=1, align='L', fill=fill)
            pdf.cell(col_widths[3], 6, res.data_nascimento.strftime('%d/%m/%Y'), border=1, align='C', fill=fill)
            pdf.cell(col_widths[4], 6, res.data_avaliacao.strftime('%d/%m/%Y'), border=1, align='C', fill=fill)
            pdf.cell(col_widths[5], 6, f"{res.peso_kg:.1f}", border=1, align='C', fill=fill)
            pdf.cell(col_widths[6], 6, encode_for_latin1(get_classification_from_indicators(res.indicadores, "Peso-para-Idade")), border=1, align='L', fill=fill)
            pdf.cell(col_widths[7], 6, f"{res.altura_cm:.1f}", border=1, align='C', fill=fill)
            pdf.cell(col_widths[8], 6, encode_for_latin1(get_classification_from_indicators(res.indicadores, "Altura-para-Idade")), border=1, align='L', fill=fill)
            pdf.cell(col_widths[9], 6, f"{res.imc:.2f}" if res.imc else "N/A", border=1, align='C', fill=fill)
            pdf.cell(col_widths[10], 6, encode_for_latin1(imc_i_class), border=1, align='L', fill=fill)
            pdf.ln()

        if report_data.batch_results.errors:
            pdf.set_font('Helvetica', 'B', 8)
            pdf.set_fill_color(254, 226, 226)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(sum(col_widths), 8, 'Registros com Erro no Processamento', border=1, align='C', fill=True)
            pdf.ln()
            pdf.set_font('Helvetica', '', 7)
            for err in report_data.batch_results.errors:
                nome = err.dados_originais.get('nome', 'N/A')
                erro_msg = f"Linha {err.linha}: {nome} - Erro: {err.erro}"
                pdf.multi_cell(sum(col_widths), 6, encode_for_latin1(erro_msg), border=1, align=Align.L, fill=True)

        # Correção: pdf.output() retorna bytearray, converter para bytes
        pdf_output = pdf.output()
        if isinstance(pdf_output, bytearray):
            pdf_bytes = bytes(pdf_output)
        else:
            pdf_bytes = pdf_output
        
        file_name = f"Relatorio_{report_data.identifier.replace(' ', '_')}.pdf"
        return Response(content=pdf_bytes, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=\"{file_name}\""})

    except Exception as e:
        print(f"Erro ao gerar PDF com fpdf2: {e}")
        raise HTTPException(status_code=500, detail="Ocorreu um erro ao gerar o relatório em PDF.")


class ExportRequest(BaseModel):
    """Modelo para requisição de exportação"""
    results: List[ResultadoProcessamentoIndividual]
    summary: dict
    escola: Optional[str] = None
    turma: Optional[str] = None


@app.post("/api/export/csv")
async def export_csv(request: ExportRequest):
    """Exporta os resultados para formato CSV"""
    try:
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Cabeçalho
        headers = [
            'ID', 'Nome', 'Sexo', 'Data Nascimento', 'Data Avaliação', 'Idade', 
            'Peso (kg)', 'Altura (cm)', 'IMC', 'Peso/Idade', 'Altura/Idade', 'IMC/Idade'
        ]
        
        if request.escola:
            headers.insert(0, 'Escola')
        if request.turma:
            headers.insert(1 if request.escola else 0, 'Turma')
            
        writer.writerow(headers)
        
        # Dados
        for resultado in request.results:
            row = []
            
            if request.escola:
                row.append(request.escola)
            if request.turma:
                row.append(request.turma)
                
            # Buscar indicadores
            peso_idade = next((ind.classificacao for ind in resultado.indicadores if 'Peso-para-Idade' in ind.tipo), 'N/A')
            altura_idade = next((ind.classificacao for ind in resultado.indicadores if 'Altura-para-Idade' in ind.tipo), 'N/A')
            imc_idade = next((ind.classificacao for ind in resultado.indicadores if 'IMC-para-Idade' in ind.tipo), 'N/A')
            
            row.extend([
                resultado.id_paciente or '',  # ID do paciente
                resultado.nome,
                resultado.sexo,
                resultado.data_nascimento.strftime('%d/%m/%Y'),
                resultado.data_avaliacao.strftime('%d/%m/%Y'),
                resultado.idade,
                float(resultado.peso_kg),
                float(resultado.altura_cm),
                resultado.imc or 'N/A',
                peso_idade,
                altura_idade,
                imc_idade
            ])
            
            writer.writerow(row)
        
        csv_content = output.getvalue()
        output.close()
        
        return Response(
            content=csv_content.encode('utf-8'),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=resultados_antropometria.csv"}
        )
        
    except Exception as e:
        print(f"Erro ao exportar CSV: {e}")
        raise HTTPException(status_code=500, detail="Erro ao gerar arquivo CSV")


@app.post("/api/export/xlsx")
async def export_xlsx(request: ExportRequest):
    """Exporta os resultados para formato Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import io
        
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Resultados Antropometria")
        else:
            ws.title = "Resultados Antropometria"
        
        # Cabeçalho com estilo
        headers = [
            'ID', 'Nome', 'Sexo', 'Data Nascimento', 'Data Avaliação', 'Idade', 
            'Peso (kg)', 'Altura (cm)', 'IMC', 'Peso/Idade', 'Altura/Idade', 'IMC/Idade'
        ]
        
        col_offset = 0
        if request.escola:
            headers.insert(0, 'Escola')
            col_offset += 1
        if request.turma:
            headers.insert(1 if request.escola else 0, 'Turma')
            col_offset += 1
        
        # Estilo do cabeçalho
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Dados
        for row_idx, resultado in enumerate(request.results, 2):
            col = 1
            
            if request.escola:
                ws.cell(row=row_idx, column=col, value=request.escola)
                col += 1
            if request.turma:
                ws.cell(row=row_idx, column=col, value=request.turma)
                col += 1
            
            # Buscar indicadores
            peso_idade = next((ind.classificacao for ind in resultado.indicadores if 'Peso-para-Idade' in ind.tipo), 'N/A')
            altura_idade = next((ind.classificacao for ind in resultado.indicadores if 'Altura-para-Idade' in ind.tipo), 'N/A')
            imc_idade = next((ind.classificacao for ind in resultado.indicadores if 'IMC-para-Idade' in ind.tipo), 'N/A')
            
            values = [
                resultado.id_paciente or '',  # ID do paciente
                resultado.nome,
                resultado.sexo,
                resultado.data_nascimento.strftime('%d/%m/%Y'),
                resultado.data_avaliacao.strftime('%d/%m/%Y'),
                resultado.idade,
                float(resultado.peso_kg),
                float(resultado.altura_cm),
                resultado.imc or 'N/A',
                peso_idade,
                altura_idade,
                imc_idade
            ]
            
            for val in values:
                ws.cell(row=row_idx, column=col, value=val)
                col += 1
        
        # Ajustar largura das colunas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Salvar em buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=resultados_antropometria.xlsx"}
        )
        
    except Exception as e:
        print(f"Erro ao exportar Excel: {e}")
        raise HTTPException(status_code=500, detail="Erro ao gerar arquivo Excel")

class ManualBatchRequest(BaseModel):
    pessoas: List[IndividuoCreate]
    identifier: str = "Lote Manual"

@app.post("/api/processar/manual-batch", response_model=BatchProcessingResponse)
async def processar_dados_manual_batch(
    request: ManualBatchRequest,
    db: Session = Depends(get_db)
):
    """Processa uma lista de pessoas enviadas manualmente e retorna os resultados."""
    try:
        service = AnthropometryService(db=db)
        
        # Converte lista de pessoas para formato similar ao CSV
        resultados_individuais = []
        erros_por_linha = []
        
        for index, pessoa in enumerate(request.pessoas, start=1):
            try:
                resultado = service.process_individual_data(pessoa)
                resultados_individuais.append(resultado)
            except Exception as e:
                erro = ErroLinha(
                    linha=index,
                    erro=str(e),
                    dados_originais={
                        "id_paciente": pessoa.id_paciente,
                        "nome": pessoa.nome,
                        "data_nascimento": pessoa.data_nascimento.isoformat() if pessoa.data_nascimento else None,
                        "data_avaliacao": pessoa.data_avaliacao.isoformat() if pessoa.data_avaliacao else None,
                        "sexo": pessoa.sexo,
                        "peso_kg": pessoa.peso_kg,
                        "altura_cm": pessoa.altura_cm
                    }
                )
                erros_por_linha.append(erro)
        
        return BatchProcessingResponse(
            success=True,
            summary={
                "total_processed": len(request.pessoas),
                "success_count": len(resultados_individuais),
                "error_count": len(erros_por_linha)
            },
            results=resultados_individuais,
            errors=erros_por_linha
        )
        
    except Exception as e:
        print(f"Erro inesperado no processamento manual batch: {e}")
        raise HTTPException(status_code=500, detail=f"Erro interno no servidor: {e}")